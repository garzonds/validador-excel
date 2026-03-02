import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL   = PatternFill("solid", start_color="2F5496")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
RESULT_PASS   = PatternFill("solid", start_color="C6EFCE")   # verde
RESULT_FAIL   = PatternFill("solid", start_color="FFCCCC")   # rojo
SUMMARY_FILL  = PatternFill("solid", start_color="D9E1F2")
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def _style_header(ws, n_cols):
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22

def generate_report(df: pd.DataFrame, out_col: str, output_path: str):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Hoja Detalle ──────────────────────────────────────────────
        df.to_excel(writer, sheet_name="Detalle", index=False)

        # ── Hoja Resumen ──────────────────────────────────────────────
        summary = (
            df[out_col]
            .value_counts()
            .reset_index()
            .rename(columns={"index": "Resultado", out_col: "Cantidad"})
        )
        summary.columns = ["Resultado", "Cantidad"]
        summary["Porcentaje"] = (summary["Cantidad"] / len(df) * 100).round(2).astype(str) + "%"
        summary.to_excel(writer, sheet_name="Resumen", index=False)

    # ── Formato con openpyxl ──────────────────────────────────────────
    wb = load_workbook(output_path)

    # Detalle
    ws_d = wb["Detalle"]
    _style_header(ws_d, df.shape[1])
    result_col_idx = df.columns.get_loc(out_col) + 1
    unique_results = df[out_col].unique().tolist()
    first_result   = unique_results[0] if unique_results else None

    for row in ws_d.iter_rows(min_row=2, max_row=ws_d.max_row):
        for cell in row:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(name="Arial", size=10)
        result_cell = row[result_col_idx - 1]
        val = result_cell.value
        result_cell.fill = RESULT_PASS if val == first_result else RESULT_FAIL
    _auto_width(ws_d)
    ws_d.freeze_panes = "A2"

    # Resumen
    ws_s = wb["Resumen"]
    _style_header(ws_s, 3)
    for row in ws_s.iter_rows(min_row=2, max_row=ws_s.max_row):
        for cell in row:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = SUMMARY_FILL
    _auto_width(ws_s)

    wb.save(output_path)
    print(f"✅ Reporte generado: {output_path}")
